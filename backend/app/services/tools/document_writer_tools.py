"""Doküman yazma tool'lari (Sprint 3.5): PDF + XLSX olusturma."""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Dict, List

from app.services.tools.base import BaseTool, ToolContext, ToolResult

logger = logging.getLogger(__name__)


class PDFGenerateTool(BaseTool):
    name = "pdf_generate"
    description = (
        "Verilen baslik + metni PDF olarak kaydet. reportlab kuruluysa onunla, "
        "yoksa fpdf2 ile dener; ikisi de yoksa hata verir."
    )
    permission = "file_system"
    parameters = {
        "type": "object",
        "properties": {
            "output_path": {"type": "string", "description": "Cikti PDF yolu"},
            "title": {"type": "string", "description": "Belge basligi"},
            "content": {"type": "string", "description": "Govde metni (cok satirli)"},
            "font_size": {"type": "integer", "description": "Yazi boyutu (varsayilan 11)"}},
        "required": ["output_path", "content"]}

    async def execute(self, args: Dict[str, Any], context: ToolContext) -> ToolResult:
        output_path = str(args.get("output_path") or "")
        content = str(args.get("content") or "")
        title = str(args.get("title") or "Belge")
        font_size = int(args.get("font_size", 11))
        if not output_path or not content.strip():
            return ToolResult(ok=False, error="output_path ve content zorunlu")

        out_p = Path(output_path).expanduser().resolve()
        out_p.parent.mkdir(parents=True, exist_ok=True)

        # Once reportlab dene
        try:
            from reportlab.lib.pagesizes import A4  # type: ignore
            from reportlab.lib.styles import getSampleStyleSheet  # type: ignore
            from reportlab.platypus import (  # type: ignore
                Paragraph,
                SimpleDocTemplate,
                Spacer,
            )
            doc = SimpleDocTemplate(str(out_p), pagesize=A4)
            styles = getSampleStyleSheet()
            elements: List[Any] = [
                Paragraph(title, styles["Heading1"]),
                Spacer(1, 12)]
            for para in content.split("\n\n"):
                # Yeni satirlari <br/> ile koru
                escaped = (
                    para.replace("&", "&amp;")
                        .replace("<", "&lt;")
                        .replace(">", "&gt;")
                        .replace("\n", "<br/>")
                )
                elements.append(Paragraph(escaped, styles["BodyText"]))
                elements.append(Spacer(1, 6))
            doc.build(elements)
            return ToolResult(
                ok=True,
                output=f"PDF olusturuldu: {out_p}",
                data={"path": str(out_p), "engine": "reportlab"},
            )
        except ImportError:
            pass
        except Exception as exc:
            logger.warning("reportlab hatasi, fpdf2'ye dusuyor: %s", exc)

        # fpdf2 fallback
        try:
            from fpdf import FPDF  # type: ignore
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Helvetica", "B", 16)
            pdf.cell(0, 10, title, ln=True)
            pdf.set_font("Helvetica", size=font_size)
            for line in content.splitlines():
                # Latin-1 dışı karakterleri kaybetmemek için multi_cell
                try:
                    pdf.multi_cell(0, font_size * 0.5 + 2, line)
                except Exception:
                    pdf.multi_cell(0, font_size * 0.5 + 2, line.encode("latin-1", errors="replace").decode("latin-1"))
            pdf.output(str(out_p))
            return ToolResult(
                ok=True,
                output=f"PDF olusturuldu (fpdf2): {out_p}",
                data={"path": str(out_p), "engine": "fpdf2"},
            )
        except ImportError:
            return ToolResult(
                ok=False,
                error="PDF icin reportlab veya fpdf2 kurulu olmali. pip install reportlab",
            )
        except Exception as exc:
            logger.exception("pdf_generate hata")
            return ToolResult(ok=False, error=f"PDF hata: {exc}")


class XLSXWriteTool(BaseTool):
    name = "xlsx_write"
    description = (
        "Bir dictionary'yi (sheets) Excel dosyasi olarak kaydeder. "
        "openpyxl gerekir."
    )
    permission = "file_system"
    parameters = {
        "type": "object",
        "properties": {
            "output_path": {"type": "string", "description": "Cikti XLSX yolu"},
            "sheets": {
                "type": "object",
                "description": "{ sheet_name: [ [row1_col1, row1_col2, ...], ... ] }",
                "additionalProperties": {
                    "type": "array",
                    "items": {"type": "array"}}}},
        "required": ["output_path", "sheets"]}

    async def execute(self, args: Dict[str, Any], context: ToolContext) -> ToolResult:
        output_path = str(args.get("output_path") or "")
        sheets = args.get("sheets") or {}
        if not output_path or not isinstance(sheets, dict) or not sheets:
            return ToolResult(ok=False, error="output_path ve sheets (dict) zorunlu")

        try:
            from openpyxl import Workbook  # type: ignore
        except ImportError:
            return ToolResult(ok=False, error="openpyxl kurulu degil. pip install openpyxl")

        out_p = Path(output_path).expanduser().resolve()
        out_p.parent.mkdir(parents=True, exist_ok=True)
        try:
            wb = Workbook()
            # Default sheet'i sil
            default = wb.active
            if default is not None:
                wb.remove(default)
            for sheet_name, rows in sheets.items():
                ws = wb.create_sheet(title=str(sheet_name)[:31])  # excel limit
                for row in rows:
                    if isinstance(row, list):
                        ws.append(row)
                    else:
                        ws.append([row])
            wb.save(str(out_p))
        except Exception as exc:
            logger.exception("xlsx_write hata")
            return ToolResult(ok=False, error=f"Excel yazma hatasi: {exc}")

        return ToolResult(
            ok=True,
            output=f"Excel olusturuldu: {out_p} ({len(sheets)} sayfa)",
            data={"path": str(out_p), "sheet_count": len(sheets)},
        )